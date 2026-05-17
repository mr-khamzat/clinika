"""
Экспорт отчётов Кабинета Директора в Excel и PDF.

Доступ — только для роли director / franchise_owner / super_admin
(см. require_director_or_owner).

Endpoints (prefix `/director/export`):
  GET /pnl.xlsx            — P&L по периодам (XLSX)
  GET /pnl.pdf             — P&L по периодам (PDF)
  GET /cashflow.xlsx       — ДДС по дням (XLSX)
  GET /cashflow.pdf        — ДДС по дням (PDF)
  GET /clinics.xlsx        — сравнение клиник (XLSX)
  GET /clinics.pdf         — сравнение клиник (PDF)
  GET /dashboard.pdf       — сводный отчёт по дашборду (PDF)

PDF делается через WeasyPrint (HTML → PDF), кириллица — fonts-dejavu,
который уже стоит в Dockerfile бэкенда.
"""
from __future__ import annotations

import html as html_lib
import io
import logging
from datetime import date, datetime, timedelta
from decimal import Decimal
from typing import Any, Optional

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.deps import require_director_or_owner
from app.database import get_db
from app.models.user import User
from app.routers.director import (
    _appointments_count,
    _default_period,
    _get_franchise_id,
    _get_tenant_ids,
    _revenue_sum,
    _top_doctors,
    _top_services,
)

logger = logging.getLogger("director_export")

router = APIRouter(prefix="/director/export", tags=["director-export"])


# ════════════════════════════════════════════════════════════════════════════
# Helpers
# ════════════════════════════════════════════════════════════════════════════


def _fmt_rub(value: Any) -> str:
    """Форматирование суммы в стиле «1 234 567 ₽» (NBSP-разделитель)."""
    try:
        v = float(value or 0)
    except (TypeError, ValueError):
        return "0 ₽"
    s = f"{v:,.0f}".replace(",", " ")
    return f"{s} ₽"


def _fmt_int(value: Any) -> str:
    try:
        v = int(value or 0)
    except (TypeError, ValueError):
        return "0"
    return f"{v:,}".replace(",", " ")


def _fmt_pct(value: Any) -> str:
    try:
        v = float(value or 0)
    except (TypeError, ValueError):
        return "0%"
    return f"{v:.1f}%"


def _fmt_date_ru(d: Any) -> str:
    if d is None:
        return ""
    if isinstance(d, str):
        return d
    try:
        return d.strftime("%d.%m.%Y")
    except Exception:  # noqa: BLE001
        return str(d)


def _xlsx_response(wb, filename: str) -> StreamingResponse:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _pdf_response(pdf_bytes: bytes, filename: str) -> StreamingResponse:
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _xlsx_apply_header(ws, row: int = 1) -> None:
    """Стиль заголовков таблицы (синий primary + белый текст)."""
    from openpyxl.styles import Alignment, Font, PatternFill  # noqa: WPS433

    for cell in ws[row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0097A7")
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _xlsx_set_widths(ws, widths: list[int]) -> None:
    """Установить ширины колонок по списку."""
    from openpyxl.utils import get_column_letter  # noqa: WPS433

    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ════════════════════════════════════════════════════════════════════════════
# Общие HTML-обёртка для PDF
# ════════════════════════════════════════════════════════════════════════════

_PDF_CSS = """
@page { size: A4; margin: 18mm 14mm 16mm 14mm; }
* { box-sizing: border-box; }
body {
  font-family: 'DejaVu Sans', sans-serif;
  font-size: 11pt; color: #111;
}
h1 { font-size: 18pt; margin: 0 0 8px 0; color: #00667a; }
h2 { font-size: 14pt; margin: 18px 0 8px 0; color: #00667a; }
.meta { color: #666; font-size: 9.5pt; margin-bottom: 14px; }
table { width: 100%; border-collapse: collapse; margin-top: 6px; }
th, td { padding: 6px 8px; border-bottom: 1px solid #ddd; font-size: 10pt; }
th {
  background: #0097A7; color: white; text-align: left;
  font-weight: 700; border: none;
}
td.num, th.num { text-align: right; font-variant-numeric: tabular-nums; }
.kpi-row { display: flex; gap: 10px; margin: 8px 0 14px 0; }
.kpi {
  flex: 1; padding: 10px 12px; border: 1px solid #e5e7eb;
  border-radius: 8px; background: #f9fafb;
}
.kpi .lbl { font-size: 9pt; color: #666; text-transform: uppercase; letter-spacing: .04em; }
.kpi .val { font-size: 14pt; font-weight: 700; color: #111; margin-top: 2px; }
.muted { color: #888; font-size: 9pt; margin-top: 10px; }
"""


def _pdf_html(title: str, meta: str, body: str) -> str:
    """Базовая HTML-обёртка для PDF-отчёта."""
    return f"""<!doctype html>
<html lang="ru"><head><meta charset="utf-8">
<title>{html_lib.escape(title)}</title>
<style>{_PDF_CSS}</style>
</head><body>
<h1>{html_lib.escape(title)}</h1>
<div class="meta">{meta}</div>
{body}
</body></html>"""


def _render_pdf(html_str: str) -> bytes:
    """Ленивый импорт WeasyPrint + рендер HTML → PDF."""
    from weasyprint import HTML  # noqa: WPS433

    return HTML(string=html_str).write_pdf()


# ════════════════════════════════════════════════════════════════════════════
# Сбор данных (повторяет логику director.py, но возвращает «голые» структуры)
# ════════════════════════════════════════════════════════════════════════════


async def _collect_pnl_series(
    db: AsyncSession,
    tenant_ids: list,
    period_from: date,
    period_to: date,
    granularity: str,
) -> list[dict[str, Any]]:
    """Серии P&L по выбранной гранулярности.

    Расходы пока не реализованы (см. /director/pnl notice) — пишем 0.
    """
    series: list[dict[str, Any]] = []
    if not tenant_ids:
        return series

    if granularity == "day":
        cursor = period_from
        while cursor <= period_to:
            start_dt = datetime.combine(cursor, datetime.min.time())
            end_dt = datetime.combine(cursor, datetime.max.time())
            rev = await _revenue_sum(db, tenant_ids, start_dt, end_dt)
            series.append(
                {
                    "label": cursor.strftime("%d.%m.%Y"),
                    "date": cursor.isoformat(),
                    "revenue": float(rev),
                    "expenses": 0.0,
                    "profit": float(rev),
                }
            )
            cursor += timedelta(days=1)
        return series

    if granularity in ("week",):
        # Неделями: ISO-неделя
        cursor = period_from
        while cursor <= period_to:
            wend = min(cursor + timedelta(days=6), period_to)
            rev = await _revenue_sum(
                db,
                tenant_ids,
                datetime.combine(cursor, datetime.min.time()),
                datetime.combine(wend, datetime.max.time()),
            )
            series.append(
                {
                    "label": (
                        f"{cursor.strftime('%d.%m')}–{wend.strftime('%d.%m.%y')}"
                    ),
                    "date": cursor.isoformat(),
                    "revenue": float(rev),
                    "expenses": 0.0,
                    "profit": float(rev),
                }
            )
            cursor = wend + timedelta(days=1)
        return series

    if granularity == "quarter":
        # Квартал = 3 месяца
        cursor = date(period_from.year, ((period_from.month - 1) // 3) * 3 + 1, 1)
        while cursor <= period_to:
            qe_month = cursor.month + 2
            qe_year = cursor.year
            if qe_month > 12:
                qe_month -= 12
                qe_year += 1
            # последний день квартала
            if qe_month == 12:
                next_q = date(qe_year + 1, 1, 1)
            else:
                next_q = date(qe_year, qe_month + 1, 1)
            qend = min(next_q - timedelta(days=1), period_to)
            qstart = max(cursor, period_from)
            rev = await _revenue_sum(
                db,
                tenant_ids,
                datetime.combine(qstart, datetime.min.time()),
                datetime.combine(qend, datetime.max.time()),
            )
            qn = (cursor.month - 1) // 3 + 1
            series.append(
                {
                    "label": f"Q{qn} {cursor.year}",
                    "date": cursor.isoformat(),
                    "revenue": float(rev),
                    "expenses": 0.0,
                    "profit": float(rev),
                }
            )
            cursor = next_q
        return series

    # default: месяцами
    cursor = date(period_from.year, period_from.month, 1)
    while cursor <= period_to:
        if cursor.month == 12:
            next_month = date(cursor.year + 1, 1, 1)
        else:
            next_month = date(cursor.year, cursor.month + 1, 1)
        mend = min(next_month - timedelta(days=1), period_to)
        mstart = max(cursor, period_from)
        rev = await _revenue_sum(
            db,
            tenant_ids,
            datetime.combine(mstart, datetime.min.time()),
            datetime.combine(mend, datetime.max.time()),
        )
        series.append(
            {
                "label": cursor.strftime("%m.%Y"),
                "date": cursor.strftime("%Y-%m"),
                "revenue": float(rev),
                "expenses": 0.0,
                "profit": float(rev),
            }
        )
        cursor = next_month
    return series


async def _collect_clinics_breakdown(
    db: AsyncSession,
    tenant_ids: list,
    period_from: date,
    period_to: date,
) -> list[dict[str, Any]]:
    """Разбивка выручки по клиникам (повторяет /director/pnl/by-clinic)."""
    from sqlalchemy import and_, func, select  # noqa: WPS433

    from app.models.clinic import Clinic  # noqa: WPS433
    from app.models.payments_clinic import (  # noqa: WPS433
        ClinicPayment,
        ClinicPaymentStatus,
    )

    if not tenant_ids:
        return []
    start_dt = datetime.combine(period_from, datetime.min.time())
    end_dt = datetime.combine(period_to, datetime.max.time())

    q = (
        select(
            ClinicPayment.clinic_id,
            func.coalesce(func.sum(ClinicPayment.amount), 0).label("revenue"),
            func.count(ClinicPayment.id).label("payments_count"),
        )
        .where(
            and_(
                ClinicPayment.tenant_id.in_(tenant_ids),
                ClinicPayment.status == ClinicPaymentStatus.SUCCEEDED,
                ClinicPayment.paid_at >= start_dt,
                ClinicPayment.paid_at <= end_dt,
            )
        )
        .group_by(ClinicPayment.clinic_id)
    )
    r = await db.execute(q)
    rows = r.all()
    clinic_ids = [row[0] for row in rows]
    rev_by_clinic = {row[0]: (row[1], row[2]) for row in rows}

    clinics_map: dict[Any, Any] = {}
    if clinic_ids:
        rc = await db.execute(select(Clinic).where(Clinic.id.in_(clinic_ids)))
        for c in rc.scalars().all():
            clinics_map[c.id] = c

    items: list[dict[str, Any]] = []
    for cid, (rev, cnt) in rev_by_clinic.items():
        c = clinics_map.get(cid)
        items.append(
            {
                "clinic_id": str(cid),
                "name": (c.name if c else "—"),
                "city": (c.city if c else ""),
                "revenue": float(rev or 0),
                "expenses": 0.0,
                "profit": float(rev or 0),
                "margin_pct": 100.0 if rev else 0.0,
                "payments_count": int(cnt or 0),
            }
        )
    items.sort(key=lambda x: x["revenue"], reverse=True)
    return items


async def _collect_cashflow_series(
    db: AsyncSession,
    tenant_ids: list,
    period_from: date,
    period_to: date,
) -> list[dict[str, Any]]:
    """ДДС по дням (повторяет /director/cashflow)."""
    from sqlalchemy import and_, func, select  # noqa: WPS433

    from app.models.payments_clinic import (  # noqa: WPS433
        ClinicPayment,
        ClinicPaymentStatus,
    )

    series: list[dict[str, Any]] = []
    if not tenant_ids:
        return series

    q = (
        select(
            func.date_trunc("day", ClinicPayment.paid_at).label("d"),
            func.coalesce(func.sum(ClinicPayment.amount), 0).label("amt"),
        )
        .where(
            and_(
                ClinicPayment.tenant_id.in_(tenant_ids),
                ClinicPayment.status == ClinicPaymentStatus.SUCCEEDED,
                ClinicPayment.paid_at
                >= datetime.combine(period_from, datetime.min.time()),
                ClinicPayment.paid_at
                <= datetime.combine(period_to, datetime.max.time()),
            )
        )
        .group_by("d")
        .order_by("d")
    )
    r = await db.execute(q)
    by_day: dict[date, Decimal] = {}
    for row in r.all():
        if row[0] is None:
            continue
        d = row[0].date() if hasattr(row[0], "date") else row[0]
        by_day[d] = Decimal(str(row[1] or 0))

    cursor = period_from
    while cursor <= period_to:
        v = float(by_day.get(cursor, Decimal("0")))
        series.append(
            {
                "date": cursor.isoformat(),
                "label": cursor.strftime("%d.%m.%Y"),
                "inflow": v,
                "outflow": 0.0,
                "net": v,
            }
        )
        cursor += timedelta(days=1)
    return series


# ════════════════════════════════════════════════════════════════════════════
# P&L  ─  XLSX
# ════════════════════════════════════════════════════════════════════════════


@router.get("/pnl.xlsx", dependencies=[Depends(require_director_or_owner)])
async def export_pnl_xlsx(
    from_: Optional[date] = Query(None, alias="from"),
    to: Optional[date] = Query(None),
    granularity: str = Query("month", regex="^(day|week|month|quarter)$"),
    user: User = Depends(require_director_or_owner),
    db: AsyncSession = Depends(get_db),
):
    """Excel-отчёт P&L (доходы/расходы/прибыль) по периодам."""
    from openpyxl import Workbook  # noqa: WPS433
    from openpyxl.styles import Alignment, Font  # noqa: WPS433

    fid = await _get_franchise_id(db, user)
    tids = await _get_tenant_ids(db, fid)
    period_from, period_to = _default_period(from_, to)
    series = await _collect_pnl_series(
        db, tids, period_from, period_to, granularity
    )
    total_rev = sum(s["revenue"] for s in series)
    total_exp = sum(s["expenses"] for s in series)
    total_prof = sum(s["profit"] for s in series)

    wb = Workbook()
    ws = wb.active
    ws.title = "P&L"

    ws.cell(1, 1, "Отчёт о доходах и расходах").font = Font(bold=True, size=14)
    ws.cell(2, 1, f"Период: {_fmt_date_ru(period_from)} – {_fmt_date_ru(period_to)}")
    ws.cell(
        3, 1, f"Сформирован: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    )
    ws.cell(4, 1, f"Гранулярность: {granularity}")

    # Шапка
    ws.cell(6, 1, "Период")
    ws.cell(6, 2, "Доходы, ₽")
    ws.cell(6, 3, "Расходы, ₽")
    ws.cell(6, 4, "Прибыль, ₽")
    _xlsx_apply_header(ws, 6)

    for i, s in enumerate(series, start=7):
        ws.cell(i, 1, s["label"])
        ws.cell(i, 2, s["revenue"]).number_format = "#,##0"
        ws.cell(i, 3, s["expenses"]).number_format = "#,##0"
        ws.cell(i, 4, s["profit"]).number_format = "#,##0"
        for c in range(2, 5):
            ws.cell(i, c).alignment = Alignment(horizontal="right")

    # Итоги
    total_row = 7 + len(series)
    ws.cell(total_row, 1, "ИТОГО").font = Font(bold=True)
    ws.cell(total_row, 2, total_rev).number_format = "#,##0"
    ws.cell(total_row, 2).font = Font(bold=True)
    ws.cell(total_row, 3, total_exp).number_format = "#,##0"
    ws.cell(total_row, 3).font = Font(bold=True)
    ws.cell(total_row, 4, total_prof).number_format = "#,##0"
    ws.cell(total_row, 4).font = Font(bold=True)

    _xlsx_set_widths(ws, [22, 18, 18, 18])

    filename = f"pnl_{period_from}_{period_to}.xlsx"
    return _xlsx_response(wb, filename)


# ════════════════════════════════════════════════════════════════════════════
# P&L  ─  PDF
# ════════════════════════════════════════════════════════════════════════════


@router.get("/pnl.pdf", dependencies=[Depends(require_director_or_owner)])
async def export_pnl_pdf(
    from_: Optional[date] = Query(None, alias="from"),
    to: Optional[date] = Query(None),
    granularity: str = Query("month", regex="^(day|week|month|quarter)$"),
    user: User = Depends(require_director_or_owner),
    db: AsyncSession = Depends(get_db),
):
    """PDF-отчёт P&L (HTML→PDF через WeasyPrint)."""
    fid = await _get_franchise_id(db, user)
    tids = await _get_tenant_ids(db, fid)
    period_from, period_to = _default_period(from_, to)
    series = await _collect_pnl_series(
        db, tids, period_from, period_to, granularity
    )
    total_rev = sum(s["revenue"] for s in series)
    total_exp = sum(s["expenses"] for s in series)
    total_prof = sum(s["profit"] for s in series)

    rows_html_parts = []
    for s in series:
        rows_html_parts.append(
            "<tr>"
            f"<td>{html_lib.escape(s['label'])}</td>"
            f"<td class='num'>{_fmt_rub(s['revenue'])}</td>"
            f"<td class='num'>{_fmt_rub(s['expenses'])}</td>"
            f"<td class='num'>{_fmt_rub(s['profit'])}</td>"
            "</tr>"
        )
    rows_html = "".join(rows_html_parts)

    body = f"""
<div class="kpi-row">
  <div class="kpi"><div class="lbl">Доходы</div><div class="val">{_fmt_rub(total_rev)}</div></div>
  <div class="kpi"><div class="lbl">Расходы</div><div class="val">{_fmt_rub(total_exp)}</div></div>
  <div class="kpi"><div class="lbl">Прибыль</div><div class="val">{_fmt_rub(total_prof)}</div></div>
</div>
<h2>Серии данных ({html_lib.escape(granularity)})</h2>
<table>
  <thead><tr>
    <th>Период</th><th class='num'>Доходы</th>
    <th class='num'>Расходы</th><th class='num'>Прибыль</th>
  </tr></thead>
  <tbody>{rows_html}</tbody>
</table>
<div class="muted">Расходы временно показываются как 0 — модуль расходов в разработке.</div>
"""
    meta = (
        f"Период: {_fmt_date_ru(period_from)} – {_fmt_date_ru(period_to)} · "
        f"Сформирован: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    )
    html_str = _pdf_html("Отчёт о доходах и расходах (P&L)", meta, body)
    pdf_bytes = _render_pdf(html_str)
    filename = f"pnl_{period_from}_{period_to}.pdf"
    return _pdf_response(pdf_bytes, filename)


# ════════════════════════════════════════════════════════════════════════════
# Cashflow  ─  XLSX + PDF
# ════════════════════════════════════════════════════════════════════════════


@router.get("/cashflow.xlsx", dependencies=[Depends(require_director_or_owner)])
async def export_cashflow_xlsx(
    from_: Optional[date] = Query(None, alias="from"),
    to: Optional[date] = Query(None),
    user: User = Depends(require_director_or_owner),
    db: AsyncSession = Depends(get_db),
):
    """Excel: ДДС по дням."""
    from openpyxl import Workbook  # noqa: WPS433
    from openpyxl.styles import Alignment, Font  # noqa: WPS433

    fid = await _get_franchise_id(db, user)
    tids = await _get_tenant_ids(db, fid)
    period_from, period_to = _default_period(from_, to)
    series = await _collect_cashflow_series(db, tids, period_from, period_to)
    total_in = sum(s["inflow"] for s in series)
    total_out = sum(s["outflow"] for s in series)

    wb = Workbook()
    ws = wb.active
    ws.title = "ДДС"

    ws.cell(1, 1, "Отчёт о движении денежных средств").font = Font(
        bold=True, size=14
    )
    ws.cell(2, 1, f"Период: {_fmt_date_ru(period_from)} – {_fmt_date_ru(period_to)}")
    ws.cell(3, 1, f"Сформирован: {datetime.now().strftime('%d.%m.%Y %H:%M')}")

    ws.cell(5, 1, "Дата")
    ws.cell(5, 2, "Приходы, ₽")
    ws.cell(5, 3, "Расходы, ₽")
    ws.cell(5, 4, "Нетто, ₽")
    _xlsx_apply_header(ws, 5)

    for i, s in enumerate(series, start=6):
        ws.cell(i, 1, s["label"])
        ws.cell(i, 2, s["inflow"]).number_format = "#,##0"
        ws.cell(i, 3, s["outflow"]).number_format = "#,##0"
        ws.cell(i, 4, s["net"]).number_format = "#,##0"
        for c in range(2, 5):
            ws.cell(i, c).alignment = Alignment(horizontal="right")

    total_row = 6 + len(series)
    ws.cell(total_row, 1, "ИТОГО").font = Font(bold=True)
    ws.cell(total_row, 2, total_in).font = Font(bold=True)
    ws.cell(total_row, 2).number_format = "#,##0"
    ws.cell(total_row, 3, total_out).font = Font(bold=True)
    ws.cell(total_row, 3).number_format = "#,##0"
    ws.cell(total_row, 4, total_in - total_out).font = Font(bold=True)
    ws.cell(total_row, 4).number_format = "#,##0"

    _xlsx_set_widths(ws, [18, 18, 18, 18])
    filename = f"cashflow_{period_from}_{period_to}.xlsx"
    return _xlsx_response(wb, filename)


@router.get("/cashflow.pdf", dependencies=[Depends(require_director_or_owner)])
async def export_cashflow_pdf(
    from_: Optional[date] = Query(None, alias="from"),
    to: Optional[date] = Query(None),
    user: User = Depends(require_director_or_owner),
    db: AsyncSession = Depends(get_db),
):
    """PDF: ДДС по дням."""
    fid = await _get_franchise_id(db, user)
    tids = await _get_tenant_ids(db, fid)
    period_from, period_to = _default_period(from_, to)
    series = await _collect_cashflow_series(db, tids, period_from, period_to)
    total_in = sum(s["inflow"] for s in series)
    total_out = sum(s["outflow"] for s in series)
    total_net = total_in - total_out

    rows_html_parts = []
    for s in series:
        rows_html_parts.append(
            "<tr>"
            f"<td>{html_lib.escape(s['label'])}</td>"
            f"<td class='num'>{_fmt_rub(s['inflow'])}</td>"
            f"<td class='num'>{_fmt_rub(s['outflow'])}</td>"
            f"<td class='num'>{_fmt_rub(s['net'])}</td>"
            "</tr>"
        )
    rows_html = "".join(rows_html_parts)

    body = f"""
<div class="kpi-row">
  <div class="kpi"><div class="lbl">Приходы</div><div class="val">{_fmt_rub(total_in)}</div></div>
  <div class="kpi"><div class="lbl">Расходы</div><div class="val">{_fmt_rub(total_out)}</div></div>
  <div class="kpi"><div class="lbl">Нетто</div><div class="val">{_fmt_rub(total_net)}</div></div>
</div>
<h2>По дням</h2>
<table>
  <thead><tr>
    <th>Дата</th><th class='num'>Приходы</th>
    <th class='num'>Расходы</th><th class='num'>Нетто</th>
  </tr></thead>
  <tbody>{rows_html}</tbody>
</table>
<div class="muted">Расходы временно показываются как 0 — модуль расходов в разработке.</div>
"""
    meta = (
        f"Период: {_fmt_date_ru(period_from)} – {_fmt_date_ru(period_to)} · "
        f"Сформирован: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    )
    html_str = _pdf_html("Движение денежных средств", meta, body)
    pdf_bytes = _render_pdf(html_str)
    filename = f"cashflow_{period_from}_{period_to}.pdf"
    return _pdf_response(pdf_bytes, filename)


# ════════════════════════════════════════════════════════════════════════════
# Clinics  ─  XLSX + PDF
# ════════════════════════════════════════════════════════════════════════════


@router.get("/clinics.xlsx", dependencies=[Depends(require_director_or_owner)])
async def export_clinics_xlsx(
    from_: Optional[date] = Query(None, alias="from"),
    to: Optional[date] = Query(None),
    user: User = Depends(require_director_or_owner),
    db: AsyncSession = Depends(get_db),
):
    """Excel: сравнение клиник."""
    from openpyxl import Workbook  # noqa: WPS433
    from openpyxl.styles import Alignment, Font  # noqa: WPS433

    fid = await _get_franchise_id(db, user)
    tids = await _get_tenant_ids(db, fid)
    period_from, period_to = _default_period(from_, to)
    items = await _collect_clinics_breakdown(db, tids, period_from, period_to)

    wb = Workbook()
    ws = wb.active
    ws.title = "Клиники"

    ws.cell(1, 1, "Сравнение клиник сети").font = Font(bold=True, size=14)
    ws.cell(2, 1, f"Период: {_fmt_date_ru(period_from)} – {_fmt_date_ru(period_to)}")
    ws.cell(3, 1, f"Сформирован: {datetime.now().strftime('%d.%m.%Y %H:%M')}")

    ws.cell(5, 1, "Клиника")
    ws.cell(5, 2, "Город")
    ws.cell(5, 3, "Выручка, ₽")
    ws.cell(5, 4, "Прибыль, ₽")
    ws.cell(5, 5, "Маржа, %")
    ws.cell(5, 6, "Платежей")
    _xlsx_apply_header(ws, 5)

    for i, c in enumerate(items, start=6):
        ws.cell(i, 1, c["name"])
        ws.cell(i, 2, c.get("city") or "")
        ws.cell(i, 3, c["revenue"]).number_format = "#,##0"
        ws.cell(i, 4, c["profit"]).number_format = "#,##0"
        ws.cell(i, 5, c["margin_pct"]).number_format = "0.0"
        ws.cell(i, 6, c["payments_count"]).number_format = "#,##0"
        for col in (3, 4, 5, 6):
            ws.cell(i, col).alignment = Alignment(horizontal="right")

    total_row = 6 + len(items)
    total_rev = sum(c["revenue"] for c in items)
    total_prof = sum(c["profit"] for c in items)
    total_cnt = sum(c["payments_count"] for c in items)
    ws.cell(total_row, 1, "ИТОГО").font = Font(bold=True)
    ws.cell(total_row, 3, total_rev).number_format = "#,##0"
    ws.cell(total_row, 3).font = Font(bold=True)
    ws.cell(total_row, 4, total_prof).number_format = "#,##0"
    ws.cell(total_row, 4).font = Font(bold=True)
    ws.cell(total_row, 6, total_cnt).number_format = "#,##0"
    ws.cell(total_row, 6).font = Font(bold=True)

    _xlsx_set_widths(ws, [32, 18, 18, 18, 12, 14])

    filename = f"clinics_{period_from}_{period_to}.xlsx"
    return _xlsx_response(wb, filename)


@router.get("/clinics.pdf", dependencies=[Depends(require_director_or_owner)])
async def export_clinics_pdf(
    from_: Optional[date] = Query(None, alias="from"),
    to: Optional[date] = Query(None),
    user: User = Depends(require_director_or_owner),
    db: AsyncSession = Depends(get_db),
):
    """PDF: сравнение клиник."""
    fid = await _get_franchise_id(db, user)
    tids = await _get_tenant_ids(db, fid)
    period_from, period_to = _default_period(from_, to)
    items = await _collect_clinics_breakdown(db, tids, period_from, period_to)

    rows_html_parts = []
    for c in items:
        rows_html_parts.append(
            "<tr>"
            f"<td>{html_lib.escape(c['name'])}</td>"
            f"<td>{html_lib.escape(c.get('city') or '')}</td>"
            f"<td class='num'>{_fmt_rub(c['revenue'])}</td>"
            f"<td class='num'>{_fmt_rub(c['profit'])}</td>"
            f"<td class='num'>{_fmt_pct(c['margin_pct'])}</td>"
            f"<td class='num'>{_fmt_int(c['payments_count'])}</td>"
            "</tr>"
        )
    rows_html = "".join(rows_html_parts) or (
        "<tr><td colspan='6' style='text-align:center;color:#888'>"
        "Нет данных за период</td></tr>"
    )

    total_rev = sum(c["revenue"] for c in items)
    total_prof = sum(c["profit"] for c in items)

    body = f"""
<div class="kpi-row">
  <div class="kpi"><div class="lbl">Клиник</div><div class="val">{_fmt_int(len(items))}</div></div>
  <div class="kpi"><div class="lbl">Выручка</div><div class="val">{_fmt_rub(total_rev)}</div></div>
  <div class="kpi"><div class="lbl">Прибыль</div><div class="val">{_fmt_rub(total_prof)}</div></div>
</div>
<table>
  <thead><tr>
    <th>Клиника</th><th>Город</th>
    <th class='num'>Выручка</th><th class='num'>Прибыль</th>
    <th class='num'>Маржа</th><th class='num'>Платежей</th>
  </tr></thead>
  <tbody>{rows_html}</tbody>
</table>
"""
    meta = (
        f"Период: {_fmt_date_ru(period_from)} – {_fmt_date_ru(period_to)} · "
        f"Сформирован: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    )
    html_str = _pdf_html("Сравнение клиник сети", meta, body)
    pdf_bytes = _render_pdf(html_str)
    filename = f"clinics_{period_from}_{period_to}.pdf"
    return _pdf_response(pdf_bytes, filename)


# ════════════════════════════════════════════════════════════════════════════
# Dashboard  ─  PDF (сводный)
# ════════════════════════════════════════════════════════════════════════════


@router.get("/dashboard.pdf", dependencies=[Depends(require_director_or_owner)])
async def export_dashboard_pdf(
    from_: Optional[date] = Query(None, alias="from"),
    to: Optional[date] = Query(None),
    user: User = Depends(require_director_or_owner),
    db: AsyncSession = Depends(get_db),
):
    """Сводный PDF-отчёт по дашборду (KPI + топ клиник/услуг/врачей)."""
    fid = await _get_franchise_id(db, user)
    tids = await _get_tenant_ids(db, fid)
    period_from, period_to = _default_period(from_, to)
    start_dt = datetime.combine(period_from, datetime.min.time())
    end_dt = datetime.combine(period_to, datetime.max.time())

    revenue = float(await _revenue_sum(db, tids, start_dt, end_dt))
    appts = await _appointments_count(db, tids, period_from, period_to)
    clinics = await _collect_clinics_breakdown(db, tids, period_from, period_to)
    services = await _top_services(db, tids, start_dt, end_dt, limit=10)
    doctors = await _top_doctors(db, tids, start_dt, end_dt, limit=10)

    # Топ-таблица клиник
    cl_rows = "".join(
        "<tr>"
        f"<td>{html_lib.escape(c['name'])}</td>"
        f"<td>{html_lib.escape(c.get('city') or '')}</td>"
        f"<td class='num'>{_fmt_rub(c['revenue'])}</td>"
        f"<td class='num'>{_fmt_int(c['payments_count'])}</td>"
        "</tr>"
        for c in clinics[:10]
    ) or "<tr><td colspan='4' style='text-align:center;color:#888'>Нет данных</td></tr>"

    sv_rows = "".join(
        "<tr>"
        f"<td>{html_lib.escape(s.get('service_name') or '—')}</td>"
        f"<td class='num'>{_fmt_rub(s.get('revenue') or 0)}</td>"
        f"<td class='num'>{_fmt_int(s.get('count') or 0)}</td>"
        "</tr>"
        for s in services
    ) or "<tr><td colspan='3' style='text-align:center;color:#888'>Нет данных</td></tr>"

    dc_rows = "".join(
        "<tr>"
        f"<td>{html_lib.escape(d.get('doctor_name') or '—')}</td>"
        f"<td>{html_lib.escape(d.get('specialty') or '')}</td>"
        f"<td class='num'>{_fmt_rub(d.get('revenue') or 0)}</td>"
        f"<td class='num'>{_fmt_int(d.get('appointments') or 0)}</td>"
        "</tr>"
        for d in doctors
    ) or "<tr><td colspan='4' style='text-align:center;color:#888'>Нет данных</td></tr>"

    body = f"""
<div class="kpi-row">
  <div class="kpi"><div class="lbl">Выручка</div><div class="val">{_fmt_rub(revenue)}</div></div>
  <div class="kpi"><div class="lbl">Приёмов</div><div class="val">{_fmt_int(appts)}</div></div>
  <div class="kpi"><div class="lbl">Клиник</div><div class="val">{_fmt_int(len(clinics))}</div></div>
</div>

<h2>Топ клиник</h2>
<table>
  <thead><tr><th>Клиника</th><th>Город</th><th class='num'>Выручка</th><th class='num'>Платежей</th></tr></thead>
  <tbody>{cl_rows}</tbody>
</table>

<h2>Топ услуг</h2>
<table>
  <thead><tr><th>Услуга</th><th class='num'>Выручка</th><th class='num'>Кол-во</th></tr></thead>
  <tbody>{sv_rows}</tbody>
</table>

<h2>Топ врачей</h2>
<table>
  <thead><tr><th>Врач</th><th>Специальность</th><th class='num'>Выработка</th><th class='num'>Приёмов</th></tr></thead>
  <tbody>{dc_rows}</tbody>
</table>
"""
    meta = (
        f"Период: {_fmt_date_ru(period_from)} – {_fmt_date_ru(period_to)} · "
        f"Сформирован: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    )
    html_str = _pdf_html("Сводный отчёт по сети", meta, body)
    pdf_bytes = _render_pdf(html_str)
    filename = f"dashboard_{period_from}_{period_to}.pdf"
    return _pdf_response(pdf_bytes, filename)
