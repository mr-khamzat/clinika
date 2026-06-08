"""
Inventory Import — импорт остатков и номенклатуры из Excel/CSV 1С (Этап 0).

Endpoints:
  POST   /inventory/import/preview        — парсит файл, возвращает превью + auto-mapping
  POST   /inventory/import/execute        — выполняет импорт по подтверждённому маппингу
  GET    /inventory/import/history        — список последних импортов
  POST   /inventory/import/{id}/rollback  — мягкий откат (deactivate items + reverse income)

Все эндпоинты тенант-изолированы и требуют активной подписки на модуль `inventory`.
Реализация переиспользует _record_movement / _require_tenant из inventory.py.
"""
import csv
import hashlib
import io
import json
import uuid
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation
from typing import Any, Optional

import openpyxl
from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile, status
from pydantic import BaseModel, Field
from sqlalchemy import and_, func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.deps import get_current_user, require_manager, get_tenant_db
from app.core.tenant import require_module
from app.database import get_db
from app.models.clinic import Clinic
from app.models.inventory import (
    InventoryCategory,
    InventoryImportLog,
    InventoryItem,
    InventoryMovement,
    InventoryMovementType,
    InventoryStock,
)
from app.models.user import User
from app.routers.inventory import _record_movement, _require_tenant


router = APIRouter(
    prefix="/inventory/import",
    tags=["inventory-import"],
    dependencies=[Depends(require_module("inventory"))],
)


# ─────────────────────── Авто-маппинг колонок (RU/EN) ──────────────────────

COLUMN_MAPPING_HINTS: dict[str, list[str]] = {
    "sku": ["артикул", "код", "номенклатура.код", "код номенклатуры", "sku"],
    "name": ["наименование", "название", "номенклатура", "товар", "name"],
    "unit": ["единица измерения", "ед. изм.", "ед.изм.", "ед изм", "ед", "unit"],
    "cost_per_unit": [
        "цена", "цена закупки", "себестоимость", "стоимость единицы",
        "стоимость", "цена за единицу", "cost", "price",
    ],
    "quantity": ["количество", "остаток", "кол-во", "кол", "qty", "quantity"],
    "category": ["группа", "категория", "вид", "category"],
    "batch_number": ["партия", "серия", "batch", "lot"],
    "expiry_date": ["срок годности", "годен до", "expiry", "expiry date"],
    "vendor": ["поставщик", "производитель", "vendor", "supplier"],
    "barcode": ["штрихкод", "штрих-код", "штрих код", "barcode", "ean"],
    "external_id": ["id в 1с", "guid", "ид", "id 1c", "external id"],
}

CATEGORY_HINTS: dict[str, InventoryCategory] = {
    "расходник": InventoryCategory.CONSUMABLE,
    "расходный": InventoryCategory.CONSUMABLE,
    "consumable": InventoryCategory.CONSUMABLE,
    "оборудование": InventoryCategory.EQUIPMENT,
    "equipment": InventoryCategory.EQUIPMENT,
    "медикамент": InventoryCategory.MEDICATION,
    "лекарство": InventoryCategory.MEDICATION,
    "medication": InventoryCategory.MEDICATION,
    "реактив": InventoryCategory.REAGENT,
    "реагент": InventoryCategory.REAGENT,
    "reagent": InventoryCategory.REAGENT,
}


def _norm(h: str) -> str:
    return (h or "").strip().lower().replace(".", " ").replace("_", " ").replace("  ", " ")


def _auto_map(headers: list[str]) -> dict[str, int]:
    """Возвращает {наш_атрибут: column_index} по first-match-hint."""
    result: dict[str, int] = {}
    norm_headers = [_norm(h) for h in headers]
    for attr, hints in COLUMN_MAPPING_HINTS.items():
        for idx, nh in enumerate(norm_headers):
            if not nh:
                continue
            for hint in hints:
                if hint == nh or hint in nh:
                    if attr not in result:
                        result[attr] = idx
                    break
            if attr in result:
                break
    return result


# ─────────────────────────── Парсеры файлов ───────────────────────────────


def _parse_xlsx(content: bytes, sheet_name: Optional[str] = None) -> tuple[list[str], list[list[str]], list[str], list[str]]:
    """Возвращает (headers, all_rows_as_strings, sheet_names, warnings)."""
    warnings: list[str] = []
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheets = wb.sheetnames
    if sheet_name and sheet_name in sheets:
        ws = wb[sheet_name]
    else:
        ws = wb[sheets[0]]
        if sheet_name:
            warnings.append(f"Лист '{sheet_name}' не найден, используется '{sheets[0]}'")

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        wb.close()
        return [], [], sheets, ["Пустой файл"]
    headers = [str(c) if c is not None else "" for c in header_row]

    all_rows: list[list[str]] = []
    for row in rows_iter:
        # пропускаем полностью пустые строки
        if all(c is None or str(c).strip() == "" for c in row):
            continue
        all_rows.append([
            ("" if c is None else (c.isoformat() if isinstance(c, (date, datetime)) else str(c)))
            for c in row
        ])
    wb.close()
    return headers, all_rows, sheets, warnings


def _parse_csv(content: bytes) -> tuple[list[str], list[list[str]], list[str], list[str]]:
    warnings: list[str] = []
    text: str | None = None
    for enc in ("utf-8-sig", "utf-8", "cp1251", "windows-1251"):
        try:
            text = content.decode(enc)
            if enc != "utf-8-sig" and enc != "utf-8":
                warnings.append(f"Кодировка {enc}")
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        text = content.decode("utf-8", errors="replace")
        warnings.append("Кодировка определена с ошибками")

    # Детект разделителя
    sample = text[:4096]
    delim = ";" if sample.count(";") > sample.count(",") else ","
    reader = csv.reader(io.StringIO(text), delimiter=delim)
    rows = list(reader)
    if not rows:
        return [], [], [], warnings + ["Пустой файл"]
    headers = [str(c).strip() for c in rows[0]]
    body = [[str(c) for c in r] for r in rows[1:] if any(str(c).strip() for c in r)]
    return headers, body, [], warnings


# ─────────────────────────── Извлечение значений ──────────────────────────


def _cell(row: list[str], mapping: dict[str, int], attr: str) -> str:
    idx = mapping.get(attr)
    if idx is None or idx >= len(row):
        return ""
    return (row[idx] or "").strip()


def _parse_decimal(v: str) -> Decimal:
    if not v:
        return Decimal("0")
    cleaned = v.replace("\xa0", "").replace(" ", "").replace(",", ".")
    try:
        return Decimal(cleaned)
    except (InvalidOperation, ValueError):
        return Decimal("0")


def _parse_date(v: str) -> Optional[date]:
    if not v:
        return None
    v = v.strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(v[:len(fmt) + 4], fmt).date()
        except ValueError:
            continue
    # ISO-форма от openpyxl date.isoformat()
    try:
        return date.fromisoformat(v[:10])
    except ValueError:
        return None


def _detect_category(v: str, default: InventoryCategory) -> InventoryCategory:
    nv = v.strip().lower()
    if not nv:
        return default
    # Прямое совпадение значения enum
    try:
        return InventoryCategory(nv)
    except ValueError:
        pass
    for hint, cat in CATEGORY_HINTS.items():
        if hint in nv:
            return cat
    return default


# ─────────────────────────── Pydantic-схемы ────────────────────────────────


class PreviewResponse(BaseModel):
    file_name: str
    file_hash: str
    source: str
    sheets: list[str] = []
    headers: list[str]
    suggested_mapping: dict[str, int]
    preview_rows: list[list[str]]
    total_rows: int
    warnings: list[str] = []


class ExecuteResult(BaseModel):
    import_id: uuid.UUID
    status: str
    rows_total: int
    rows_created: int
    rows_updated: int
    rows_skipped: int
    rows_failed: int
    errors: list[dict] = []


class HistoryRow(BaseModel):
    id: uuid.UUID
    file_name: str
    source: str
    status: str
    rows_total: int
    rows_created: int
    rows_updated: int
    rows_skipped: int
    rows_failed: int
    clinic_id: Optional[uuid.UUID] = None
    created_at: datetime
    completed_at: Optional[datetime] = None


# ─────────────────────────── Endpoints ────────────────────────────────────


@router.post("/preview", response_model=PreviewResponse)
async def preview_import(
    file: UploadFile = File(...),
    sheet_name: Optional[str] = Form(None),
    user: User = Depends(require_manager),
    db: AsyncSession = Depends(get_tenant_db),
):
    """Парсит файл, возвращает заголовки + автомаппинг + первые 10 строк."""
    _require_tenant(user)
    raw = await file.read()
    if not raw:
        raise HTTPException(400, "Пустой файл")

    file_hash = hashlib.sha256(raw).hexdigest()
    fname = file.filename or "upload"
    ext = fname.lower().rsplit(".", 1)[-1] if "." in fname else ""
    if ext in ("xlsx", "xlsm", "xls"):
        source = "xlsx"
        headers, rows, sheets, warnings = _parse_xlsx(raw, sheet_name)
    elif ext == "csv":
        source = "csv"
        headers, rows, sheets, warnings = _parse_csv(raw)
    else:
        # пробуем определить по содержимому
        if raw[:2] == b"PK":
            source = "xlsx"
            headers, rows, sheets, warnings = _parse_xlsx(raw, sheet_name)
        else:
            source = "csv"
            headers, rows, sheets, warnings = _parse_csv(raw)

    if not headers:
        raise HTTPException(400, "Не удалось прочитать заголовки файла")

    suggested = _auto_map(headers)
    if "sku" not in suggested and "name" not in suggested:
        warnings.append("Не найдены колонки SKU/Наименование — требуется ручной маппинг")

    return PreviewResponse(
        file_name=fname,
        file_hash=file_hash,
        source=source,
        sheets=sheets,
        headers=headers,
        suggested_mapping=suggested,
        preview_rows=rows[:10],
        total_rows=len(rows),
        warnings=warnings,
    )


@router.post("/execute", response_model=ExecuteResult)
async def execute_import(
    file: UploadFile = File(...),
    clinic_id: uuid.UUID = Form(...),
    mapping: str = Form(..., description="JSON: {'sku':0,'name':1,...}"),
    existing_strategy: str = Form("update", description="skip|update|replace"),
    default_category: InventoryCategory = Form(InventoryCategory.CONSUMABLE),
    default_vendor: Optional[str] = Form(None),
    default_supplier_name: Optional[str] = Form(None),
    paid_at: Optional[str] = Form(None),
    sheet_name: Optional[str] = Form(None),
    user: User = Depends(require_manager),
    db: AsyncSession = Depends(get_tenant_db),
):
    """Импорт по подтверждённому маппингу. Создаёт лог и движения INCOME."""
    tenant_id = _require_tenant(user)

    if existing_strategy not in ("skip", "update", "replace"):
        raise HTTPException(400, "existing_strategy: skip|update|replace")

    try:
        mapping_dict: dict[str, int] = json.loads(mapping)
        mapping_dict = {k: int(v) for k, v in mapping_dict.items() if v is not None and v != ""}
    except (json.JSONDecodeError, ValueError, TypeError):
        raise HTTPException(400, "mapping: некорректный JSON")

    if "sku" not in mapping_dict and "name" not in mapping_dict:
        raise HTTPException(400, "Маппинг должен содержать sku или name")

    # Проверка клиники
    clinic = (
        await db.execute(
            select(Clinic).where(Clinic.id == clinic_id, Clinic.tenant_id == tenant_id)
        )
    ).scalar_one_or_none()
    if not clinic:
        raise HTTPException(404, "Клиника не найдена")

    raw = await file.read()
    if not raw:
        raise HTTPException(400, "Пустой файл")
    file_hash = hashlib.sha256(raw).hexdigest()
    fname = file.filename or "upload"
    ext = fname.lower().rsplit(".", 1)[-1] if "." in fname else ""
    if ext in ("xlsx", "xlsm", "xls") or raw[:2] == b"PK":
        source = "xlsx"
        _headers, rows, _sheets, _w = _parse_xlsx(raw, sheet_name)
    else:
        source = "csv"
        _headers, rows, _sheets, _w = _parse_csv(raw)

    # Создаём лог-запись pending
    log = InventoryImportLog(
        tenant_id=tenant_id,
        clinic_id=clinic_id,
        source=source,
        file_name=fname[:255],
        file_hash=file_hash,
        rows_total=len(rows),
        mapping_used=mapping_dict,
        status="pending",
        created_by_id=user.id,
    )
    db.add(log)
    await db.flush()

    rows_created = 0
    rows_updated = 0
    rows_skipped = 0
    rows_failed = 0
    errors: list[dict] = []

    for line_no, row in enumerate(rows, start=2):
        try:
            sku = _cell(row, mapping_dict, "sku")
            name = _cell(row, mapping_dict, "name")
            if not sku and not name:
                rows_skipped += 1
                continue
            if not sku:
                # генерим sku из name
                sku = f"AUTO-{hashlib.md5(name.encode()).hexdigest()[:10]}"
            if not name:
                name = sku

            unit = _cell(row, mapping_dict, "unit") or "шт"
            barcode = _cell(row, mapping_dict, "barcode") or None
            vendor = _cell(row, mapping_dict, "vendor") or default_vendor or None
            cost = _parse_decimal(_cell(row, mapping_dict, "cost_per_unit"))
            quantity = _parse_decimal(_cell(row, mapping_dict, "quantity"))
            batch = _cell(row, mapping_dict, "batch_number") or ""
            expiry = _parse_date(_cell(row, mapping_dict, "expiry_date"))
            category = _detect_category(_cell(row, mapping_dict, "category"), default_category)

            existing: Optional[InventoryItem] = (
                await db.execute(
                    select(InventoryItem).where(
                        InventoryItem.tenant_id == tenant_id,
                        InventoryItem.sku == sku,
                    )
                )
            ).scalar_one_or_none()

            if existing:
                if existing_strategy == "skip":
                    rows_skipped += 1
                    continue
                # update / replace
                existing.name = name[:200] or existing.name
                existing.unit = (unit or existing.unit)[:20]
                if barcode:
                    existing.barcode = barcode[:100]
                if vendor:
                    existing.vendor = vendor[:200]
                if cost > 0 or existing_strategy == "replace":
                    existing.cost_per_unit = cost
                if existing_strategy == "replace":
                    existing.category = category
                    if expiry:
                        existing.expiry_tracked = True
                existing.is_active = True
                item = existing
                rows_updated += 1
            else:
                item = InventoryItem(
                    tenant_id=tenant_id,
                    sku=sku[:50],
                    name=name[:200],
                    category=category,
                    unit=unit[:20],
                    barcode=barcode[:100] if barcode else None,
                    vendor=vendor[:200] if vendor else None,
                    cost_per_unit=cost,
                    expiry_tracked=bool(expiry),
                    is_active=True,
                )
                db.add(item)
                await db.flush()
                rows_created += 1

            # Зачисление прихода
            if quantity > 0:
                await _record_movement(
                    db,
                    tenant_id=tenant_id,
                    item_id=item.id,
                    clinic_id=clinic_id,
                    mtype=InventoryMovementType.INCOME,
                    delta=quantity,
                    batch=batch,
                    expiry=expiry,
                    ref_entity_type="inventory_import",
                    ref_entity_id=log.id,
                    comment=f"1С импорт: {fname}",
                    performed_by_user_id=user.id,
                )
        except HTTPException as he:
            rows_failed += 1
            errors.append({"line": line_no, "error": str(he.detail)[:300]})
        except Exception as e:
            rows_failed += 1
            errors.append({"line": line_no, "error": str(e)[:300]})

    log.rows_created = rows_created
    log.rows_updated = rows_updated
    log.rows_skipped = rows_skipped
    log.rows_failed = rows_failed
    log.errors = errors[:200] if errors else None
    log.result_summary = {
        "default_category": default_category.value,
        "default_vendor": default_vendor,
        "default_supplier_name": default_supplier_name,
        "paid_at": paid_at,
        "existing_strategy": existing_strategy,
    }
    log.status = "completed" if rows_failed == 0 else ("completed" if (rows_created + rows_updated) > 0 else "failed")
    log.completed_at = datetime.now(timezone.utc)

    await db.commit()

    return ExecuteResult(
        import_id=log.id,
        status=log.status,
        rows_total=log.rows_total,
        rows_created=rows_created,
        rows_updated=rows_updated,
        rows_skipped=rows_skipped,
        rows_failed=rows_failed,
        errors=errors[:50],
    )


@router.get("/history")
async def list_history(
    limit: int = Query(50, ge=1, le=200),
    offset: int = Query(0, ge=0),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_tenant_db),
):
    """История последних импортов тенанта."""
    tenant_id = _require_tenant(user)
    q = select(InventoryImportLog).where(InventoryImportLog.tenant_id == tenant_id)
    total = (await db.execute(select(func.count()).select_from(q.subquery()))).scalar_one()
    rows = (
        await db.execute(
            q.order_by(InventoryImportLog.created_at.desc()).limit(limit).offset(offset)
        )
    ).scalars().all()
    return {
        "total": total,
        "items": [
            HistoryRow.model_validate(r, from_attributes=True).model_dump(mode="json")
            for r in rows
        ],
    }


@router.post("/{import_id}/rollback")
async def rollback_import(
    import_id: uuid.UUID,
    user: User = Depends(require_manager),
    db: AsyncSession = Depends(get_tenant_db),
):
    """Мягкий откат: deactivate items, созданные импортом + reverse income-движения.

    Логика:
    - Берём все INCOME-движения с ref_entity_type='inventory_import' и ref_entity_id=import_id.
    - Для каждого создаём WRITE_OFF на ту же quantity (через _record_movement → delta=-qty).
    - Если у item нет ИНЫХ outgoing-движений и других импортов → is_active=False.
    - Если хотя бы по одному item уже был расход → откат всё равно делаем (но item оставляем активным).
    - Status → rolled_back.
    """
    tenant_id = _require_tenant(user)
    log = (
        await db.execute(
            select(InventoryImportLog).where(
                InventoryImportLog.id == import_id,
                InventoryImportLog.tenant_id == tenant_id,
            )
        )
    ).scalar_one_or_none()
    if not log:
        raise HTTPException(404, "Импорт не найден")
    if log.status == "rolled_back":
        raise HTTPException(409, "Импорт уже откачен")
    if log.status != "completed":
        raise HTTPException(409, f"Нельзя откатить импорт со статусом {log.status}")

    movements = (
        await db.execute(
            select(InventoryMovement).where(
                InventoryMovement.tenant_id == tenant_id,
                InventoryMovement.ref_entity_type == "inventory_import",
                InventoryMovement.ref_entity_id == import_id,
                InventoryMovement.type == InventoryMovementType.INCOME,
            )
        )
    ).scalars().all()

    reversed_count = 0
    failed: list[dict] = []
    affected_item_ids: set[uuid.UUID] = set()

    for mov in movements:
        affected_item_ids.add(mov.item_id)
        try:
            await _record_movement(
                db,
                tenant_id=tenant_id,
                item_id=mov.item_id,
                clinic_id=mov.clinic_id,
                mtype=InventoryMovementType.WRITE_OFF,
                delta=-mov.quantity,
                batch=mov.batch_number or "",
                expiry=mov.expiry_date,
                ref_entity_type="inventory_import_rollback",
                ref_entity_id=log.id,
                comment=f"Откат импорта {log.file_name}",
                performed_by_user_id=user.id,
            )
            reversed_count += 1
        except HTTPException as he:
            failed.append({"movement_id": str(mov.id), "error": str(he.detail)[:200]})
        except Exception as e:
            failed.append({"movement_id": str(mov.id), "error": str(e)[:200]})

    # Soft-deactivate items, у которых нет outgoing-движений вне этого импорта
    deactivated = 0
    for item_id in affected_item_ids:
        outgoing_cnt = (
            await db.execute(
                select(func.count(InventoryMovement.id)).where(
                    InventoryMovement.tenant_id == tenant_id,
                    InventoryMovement.item_id == item_id,
                    InventoryMovement.type.in_(
                        [
                            InventoryMovementType.OUTGOING,
                            InventoryMovementType.TRANSFER,
                        ]
                    ),
                )
            )
        ).scalar_one()
        if outgoing_cnt == 0:
            await db.execute(
                InventoryItem.__table__.update()
                .where(
                    InventoryItem.id == item_id,
                    InventoryItem.tenant_id == tenant_id,
                )
                .values(is_active=False)
            )
            deactivated += 1

    log.status = "rolled_back"
    log.result_summary = {
        **(log.result_summary or {}),
        "rollback": {
            "reversed_movements": reversed_count,
            "deactivated_items": deactivated,
            "failed": failed[:50],
            "rolled_back_at": datetime.now(timezone.utc).isoformat(),
        },
    }
    await db.commit()

    return {
        "import_id": str(log.id),
        "status": log.status,
        "reversed_movements": reversed_count,
        "deactivated_items": deactivated,
        "failed": failed[:50],
    }
