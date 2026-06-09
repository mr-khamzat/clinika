"""
LTV-аналитика — endpoints модуля ltv_pro.

  GET  /analytics/ltv/patients?clinic_id&limit=100&min_visits=2&repeat_only&inactive_days
       — топ по LTV (с фильтрами «повторные» / «спящие»)
  GET  /analytics/ltv/cohorts?period=quarter                    — когорты
  GET  /analytics/ltv/summary?clinic_id                         — сводка
  POST /analytics/ltv/recompute?clinic_id                       — принудительный пересчёт
  GET  /analytics/ltv/export/pdf?clinic_id&period               — PDF-отчёт
  GET  /analytics/ltv/export/xlsx?clinic_id                     — Excel-отчёт
  GET  /analytics/ltv/contacts.csv?clinic_id&min_visits&inactive_days&format=csv|xlsx
       — экспорт контактов пациентов (CSV UTF-8 BOM или XLSX)

Все требуют:
  - роль manager и выше (require_manager)
  - активную подписку модуля ltv_pro (require_module)

Scope-логика (фильтр по клиникам):
  - Если clinic_id НЕ передан — для пользователей с user.clinic_id (manager,
    привязанный к клинике, или reg/nurse) автоматически подставляется их
    собственный clinic_id (кроме franchise_owner — он видит все клиники).
  - Если clinic_id передан — проверяется через get_user_clinic_ids(),
    что у пользователя есть к нему доступ. Иначе 403.
"""
import csv
import io
import uuid
from datetime import datetime, timedelta
from decimal import Decimal
from typing import Optional
from urllib.parse import quote

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import Response
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.deps import require_manager, get_tenant_db
from app.core.tenant import get_current_tenant, require_module
from app.database import get_db
from app.models.ltv import PatientLtvSnapshot
from app.models.tenant import Tenant
from app.models.user import User, UserRole
from app.routers.manager.clinics_access import get_user_clinic_ids
from app.services.ltv_export_service import generate_ltv_excel, generate_ltv_pdf
from app.services.ltv_service import compute_cohorts, compute_ltv_for_clinic

router = APIRouter(prefix="/analytics/ltv", tags=["ltv"])

_mgr = Depends(require_manager)
_mod = Depends(require_module("ltv_pro"))


# ───────────────────────────────────────────────────────────────────────────
# Хелпер: разрешение clinic_id с учётом прав пользователя
# ───────────────────────────────────────────────────────────────────────────

async def _resolve_clinic_scope(
    db: AsyncSession,
    user: User,
    clinic_id: Optional[uuid.UUID],
) -> Optional[uuid.UUID]:
    """
    Возвращает фактический clinic_id для фильтрации.

    Правила:
      • franchise_owner и super_admin без clinic_id → возвращает clinic_id «как есть»
        (None → агрегат по всем клиникам тенанта/франшизы).
      • Остальные пользователи:
          - если clinic_id не передан → подставляем user.clinic_id (если задан)
          - если clinic_id передан → проверяем доступ через get_user_clinic_ids()
      • Если запрашиваемая клиника недоступна — 403.
    """
    if clinic_id is None:
        # franchise_owner / super_admin без явного фильтра — агрегат
        if user.role in (UserRole.FRANCHISE_OWNER, UserRole.SUPER_ADMIN):
            return None
        # Прочие — подставляем user.clinic_id если задан
        if user.clinic_id is not None:
            return user.clinic_id
        # manager без clinic_id → видит все клиники тенанта (None)
        return None

    # clinic_id передан → проверяем доступ
    accessible = await get_user_clinic_ids(db, user)
    # super_admin без выбранного тенанта (accessible пуст) — расширяем до tenant
    if user.role == UserRole.SUPER_ADMIN and not accessible:
        if user.tenant_id is None:
            raise HTTPException(status_code=403, detail="Тенант не выбран")
        accessible = await get_user_clinic_ids(db, user, tenant_id_param=user.tenant_id)
    if clinic_id not in accessible:
        raise HTTPException(status_code=403, detail="Нет доступа к этой клинике")
    return clinic_id


def _days_since(dt: Optional[datetime]) -> Optional[int]:
    """Сколько целых дней прошло с момента dt до now (UTC). None если dt пуст."""
    if dt is None:
        return None
    delta = datetime.utcnow() - dt
    # отрицательные значения (визит «в будущем» из-за TZ) клампим к 0
    return max(0, int(delta.days))


# ───────────────────────────────────────────────────────────────────────────
# Горизонт расчёта LTV (1/3/5/10 лет)
# ───────────────────────────────────────────────────────────────────────────

# Базовый горизонт, под который считаются и сохраняются снапшоты в БД.
# Если фронт просит другой горизонт — пересчитываем на лету коэффициентом.
_BASE_LTV_HORIZON_YEARS = Decimal("3")


def _clamp_years(years: Optional[int]) -> int:
    """Ограничивает years диапазоном 1..10. None → 3."""
    if years is None:
        return 3
    try:
        n = int(years)
    except (TypeError, ValueError):
        return 3
    if n < 1:
        return 1
    if n > 10:
        return 10
    return n


def _scale_factor(years: int) -> Decimal:
    """Коэффициент пересчёта LTV из базовых 3 лет в произвольный горизонт."""
    return (Decimal(years) / _BASE_LTV_HORIZON_YEARS)


def _rescale(value, years: int) -> float:
    """Пересчёт сохранённого LTV/NetLTV под выбранный горизонт.

    В БД лежат значения для 3-летнего горизонта, поэтому новый = старый × years / 3.
    Возвращает float (готовое к JSON). На вход — Decimal/float/None.
    """
    if value is None:
        return 0.0
    try:
        d = Decimal(str(value))
    except Exception:
        return 0.0
    res = (d * _scale_factor(years)).quantize(Decimal("0.01"))
    return float(res)


@router.get("/patients", dependencies=[_mgr, _mod])
async def list_top_patients(
    clinic_id: Optional[uuid.UUID] = Query(None, description="UUID клиники, либо все клиники тенанта"),
    limit: int = Query(100, ge=1, le=500),
    min_visits: int = Query(2, ge=1, description="Минимум визитов для попадания в выборку"),
    repeat_only: bool = Query(False, description="Только повторные (visits_count ≥ 2)"),
    inactive_days: Optional[int] = Query(
        None, ge=1, le=3650,
        description="Фильтр спящих пациентов: дней с последнего визита ≥ inactive_days",
    ),
    years: int = Query(
        3, ge=1, le=10,
        description="Горизонт расчёта LTV (1..10 лет). По умолчанию 3 — соответствует БД.",
    ),
    user: User = Depends(require_manager),
    tenant: Tenant | None = Depends(get_current_tenant),
    db: AsyncSession = Depends(get_tenant_db),
):
    """Топ пациентов по LTV (DESC). Поддержка фильтров «повторные» / «спящие».

    `years` пересчитывает ltv_estimate / net_ltv в response из БД-значений
    (которые рассчитаны под 3 года) — без переписывания снапшотов.
    """
    if tenant is None:
        return []

    years = _clamp_years(years)
    effective_clinic_id = await _resolve_clinic_scope(db, user, clinic_id)

    # Если включён repeat_only — поднимаем порог визитов до 2 (если он ниже).
    effective_min_visits = max(min_visits, 2) if repeat_only else min_visits

    q = select(PatientLtvSnapshot).where(
        PatientLtvSnapshot.tenant_id == tenant.id,
        PatientLtvSnapshot.visits_count >= effective_min_visits,
    )
    if effective_clinic_id is not None:
        q = q.where(PatientLtvSnapshot.clinic_id == effective_clinic_id)

    # Фильтр «спящие»: last_visit_at ≤ (now - inactive_days)
    # SQL уровень — чтобы не тащить лишние строки.
    if inactive_days is not None:
        threshold = datetime.utcnow() - timedelta(days=inactive_days)
        q = q.where(PatientLtvSnapshot.last_visit_at <= threshold)

    q = q.order_by(PatientLtvSnapshot.ltv_estimate.desc()).limit(limit)

    rows = (await db.execute(q)).scalars().all()
    return [
        {
            "id": str(r.id),
            "patient_phone": r.patient_phone,
            "patient_name": r.patient_name,
            "visits_count": r.visits_count,
            "total_spent": float(r.total_spent or 0),
            "avg_check": float(r.avg_check or 0),
            # ltv_estimate / net_ltv пересчитаны под выбранный horizon=years.
            "ltv_estimate": _rescale(r.ltv_estimate, years),
            # NetLTV по фактическим оплатам (getPayments). 0 — данные пока недоступны.
            "net_ltv": _rescale(r.net_ltv, years),
            "visits_per_year": float(r.visits_per_year or 0),
            "first_visit_at": r.first_visit_at.isoformat() if r.first_visit_at else None,
            "last_visit_at": r.last_visit_at.isoformat() if r.last_visit_at else None,
            "days_since_last_visit": _days_since(r.last_visit_at),
            "cohort_quarter": r.cohort_quarter,
            "churn_risk": r.churn_risk,
            "clinic_id": str(r.clinic_id) if r.clinic_id else None,
            "horizon_years": years,
        }
        for r in rows
    ]


@router.get("/cohorts", dependencies=[_mgr, _mod])
async def list_cohorts(
    period: str = Query("quarter", pattern="^(quarter)$"),
    tenant: Tenant | None = Depends(get_current_tenant),
    db: AsyncSession = Depends(get_tenant_db),
):
    """Сводка по когортам (по умолчанию — квартал первого визита)."""
    if tenant is None:
        return []
    return await compute_cohorts(db, tenant.id, period=period)


@router.get("/summary", dependencies=[_mgr, _mod])
async def get_summary(
    clinic_id: Optional[uuid.UUID] = Query(None),
    years: int = Query(
        3, ge=1, le=10,
        description="Горизонт расчёта LTV (1..10 лет). По умолчанию 3 — соответствует БД.",
    ),
    user: User = Depends(require_manager),
    tenant: Tenant | None = Depends(get_current_tenant),
    db: AsyncSession = Depends(get_tenant_db),
):
    """Общие метрики: avg LTV, avg NetLTV, total patients, churn rate, at-risk.

    `years` пересчитывает avg_ltv / avg_net_ltv в response из БД-значений
    (рассчитанных под 3 года) — снапшоты не переписываются.
    """
    if tenant is None:
        return {
            "total_patients": 0,
            "avg_ltv": 0,
            "avg_net_ltv": 0,
            "total_spent": 0,
            "avg_check": 0,
            "churn_rate": 0,
            "at_risk_patients": 0,
            "last_computed_at": None,
            "horizon_years": _clamp_years(years),
        }

    years = _clamp_years(years)
    effective_clinic_id = await _resolve_clinic_scope(db, user, clinic_id)

    base = select(
        func.count(PatientLtvSnapshot.id),
        func.coalesce(func.avg(PatientLtvSnapshot.ltv_estimate), 0),
        func.coalesce(func.sum(PatientLtvSnapshot.total_spent), 0),
        func.coalesce(func.avg(PatientLtvSnapshot.avg_check), 0),
        func.max(PatientLtvSnapshot.computed_at),
        # Средний NetLTV считаем только по тем пациентам, у кого net_ltv > 0
        # (т.к. при отсутствии getPayments значение = 0 и оно бы занижало среднее).
        func.coalesce(
            func.avg(
                func.nullif(PatientLtvSnapshot.net_ltv, 0)
            ),
            0,
        ),
    ).where(PatientLtvSnapshot.tenant_id == tenant.id)
    if effective_clinic_id is not None:
        base = base.where(PatientLtvSnapshot.clinic_id == effective_clinic_id)

    row = (await db.execute(base)).one()
    total_patients, avg_ltv, total_spent, avg_check, last_computed, avg_net_ltv = row

    at_risk_q = select(func.count(PatientLtvSnapshot.id)).where(
        PatientLtvSnapshot.tenant_id == tenant.id,
        PatientLtvSnapshot.churn_risk == "high",
    )
    if effective_clinic_id is not None:
        at_risk_q = at_risk_q.where(PatientLtvSnapshot.clinic_id == effective_clinic_id)
    at_risk = int((await db.execute(at_risk_q)).scalar() or 0)

    medium_q = select(func.count(PatientLtvSnapshot.id)).where(
        PatientLtvSnapshot.tenant_id == tenant.id,
        PatientLtvSnapshot.churn_risk == "medium",
    )
    if effective_clinic_id is not None:
        medium_q = medium_q.where(PatientLtvSnapshot.clinic_id == effective_clinic_id)
    medium = int((await db.execute(medium_q)).scalar() or 0)

    total = int(total_patients or 0)
    churn_rate = round(((at_risk + medium) / total * 100.0), 2) if total else 0.0

    return {
        "total_patients": total,
        # avg_ltv / avg_net_ltv пересчитаны под выбранный горизонт (years).
        "avg_ltv": _rescale(avg_ltv, years),
        "avg_net_ltv": _rescale(avg_net_ltv, years),
        "total_spent": float(total_spent or 0),
        "avg_check": float(avg_check or 0),
        "churn_rate": churn_rate,
        "at_risk_patients": at_risk,
        "medium_risk_patients": medium,
        "last_computed_at": last_computed.isoformat() if last_computed else None,
        "horizon_years": years,
    }


@router.post("/recompute", dependencies=[_mgr, _mod])
async def recompute(
    clinic_id: Optional[uuid.UUID] = Query(None),
    user: User = Depends(require_manager),
    tenant: Tenant | None = Depends(get_current_tenant),
    db: AsyncSession = Depends(get_tenant_db),
):
    """Принудительный пересчёт LTV-снапшотов. Возвращает {updated, patients}."""
    if tenant is None:
        raise HTTPException(status_code=400, detail="Тенант не определён")
    effective_clinic_id = await _resolve_clinic_scope(db, user, clinic_id)
    try:
        result = await compute_ltv_for_clinic(db, tenant, effective_clinic_id)
        return {"ok": True, **result}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка пересчёта: {e}")


# ───────────────────────────────────────────────────────────────────────────
# Экспорт отчётов (PDF / Excel)
# ───────────────────────────────────────────────────────────────────────────


def _content_disposition(filename: str) -> str:
    """Content-Disposition с поддержкой UTF-8 (RFC 5987) — для русских имён."""
    encoded = quote(filename)
    return f"attachment; filename=\"{encoded}\"; filename*=UTF-8''{encoded}"


@router.get("/export/pdf", dependencies=[_mgr, _mod])
async def export_pdf(
    clinic_id: Optional[uuid.UUID] = Query(None),
    period: str = Query("all", description="Метка периода для шапки: all/month/quarter/year"),
    years: int = Query(
        3, ge=1, le=10,
        description="Горизонт расчёта LTV (1..10 лет). По умолчанию 3.",
    ),
    user: User = Depends(require_manager),
    tenant: Tenant | None = Depends(get_current_tenant),
    db: AsyncSession = Depends(get_tenant_db),
):
    """PDF-отчёт LTV: KPI, топ-50 пациентов, когорты, диаграммы."""
    if tenant is None:
        raise HTTPException(status_code=400, detail="Тенант не определён")
    years = _clamp_years(years)
    effective_clinic_id = await _resolve_clinic_scope(db, user, clinic_id)
    try:
        pdf_bytes = await generate_ltv_pdf(db, tenant, effective_clinic_id, period, years=years)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка генерации PDF: {e}")

    today = datetime.utcnow().strftime("%Y-%m-%d")
    filename = f"LTV-отчёт-АРЦ-{today}.pdf"
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": _content_disposition(filename)},
    )


@router.get("/export/xlsx", dependencies=[_mgr, _mod])
async def export_xlsx(
    clinic_id: Optional[uuid.UUID] = Query(None),
    years: int = Query(
        3, ge=1, le=10,
        description="Горизонт расчёта LTV (1..10 лет). По умолчанию 3.",
    ),
    user: User = Depends(require_manager),
    tenant: Tenant | None = Depends(get_current_tenant),
    db: AsyncSession = Depends(get_tenant_db),
):
    """Excel-отчёт LTV: листы Сводка / Топ пациентов / Когорты."""
    if tenant is None:
        raise HTTPException(status_code=400, detail="Тенант не определён")
    years = _clamp_years(years)
    effective_clinic_id = await _resolve_clinic_scope(db, user, clinic_id)
    try:
        xlsx_bytes = await generate_ltv_excel(db, tenant, effective_clinic_id, years=years)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка генерации Excel: {e}")

    today = datetime.utcnow().strftime("%Y-%m-%d")
    filename = f"LTV-отчёт-АРЦ-{today}.xlsx"
    return Response(
        content=xlsx_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": _content_disposition(filename)},
    )


# ───────────────────────────────────────────────────────────────────────────
# Экспорт контактов пациентов (CSV / XLSX)
# ───────────────────────────────────────────────────────────────────────────

def _fmt_date_ru(dt: Optional[datetime]) -> str:
    """ДД.ММ.ГГГГ или пусто."""
    if dt is None:
        return ""
    return dt.strftime("%d.%m.%Y")


async def _fetch_contacts(
    db: AsyncSession,
    tenant: Tenant,
    clinic_id: Optional[uuid.UUID],
    min_visits: int,
    inactive_days: Optional[int],
) -> list[dict]:
    """Тянет пациентов под текущие фильтры, формирует строки для экспорта."""
    q = select(PatientLtvSnapshot).where(
        PatientLtvSnapshot.tenant_id == tenant.id,
        PatientLtvSnapshot.visits_count >= min_visits,
    )
    if clinic_id is not None:
        q = q.where(PatientLtvSnapshot.clinic_id == clinic_id)
    if inactive_days is not None:
        threshold = datetime.utcnow() - timedelta(days=inactive_days)
        q = q.where(PatientLtvSnapshot.last_visit_at <= threshold)
    # Сортируем по LTV, чтобы самые ценные были сверху
    q = q.order_by(PatientLtvSnapshot.ltv_estimate.desc())
    rows = (await db.execute(q)).scalars().all()

    out: list[dict] = []
    for r in rows:
        out.append({
            "phone": r.patient_phone or "",
            "name": r.patient_name or "",
            "visits": int(r.visits_count or 0),
            "first_visit": _fmt_date_ru(r.first_visit_at),
            "last_visit": _fmt_date_ru(r.last_visit_at),
            "days_since_last": _days_since(r.last_visit_at) or 0,
            "total_spent": float(r.total_spent or 0),
            "ltv": float(r.ltv_estimate or 0),
        })
    return out


def _build_contacts_csv(contacts: list[dict]) -> bytes:
    """Собирает CSV (UTF-8 BOM, разделитель «;» — для русского Excel)."""
    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=";", quoting=csv.QUOTE_MINIMAL)
    writer.writerow([
        "Телефон",
        "ФИО",
        "Визитов",
        "Первый визит",
        "Последний визит",
        "Дней с последнего визита",
        "Сумма (₽)",
        "LTV (₽)",
    ])
    for c in contacts:
        writer.writerow([
            c["phone"],
            c["name"],
            c["visits"],
            c["first_visit"],
            c["last_visit"],
            c["days_since_last"],
            f"{int(round(c['total_spent']))}",
            f"{int(round(c['ltv']))}",
        ])
    # UTF-8 BOM для корректного отображения кириллицы в Excel
    return ("﻿" + buf.getvalue()).encode("utf-8")


def _build_contacts_xlsx(contacts: list[dict]) -> bytes:
    """Собирает XLSX с теми же колонками, что и CSV."""
    from openpyxl import Workbook  # noqa: WPS433
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # noqa: WPS433

    wb = Workbook()
    ws = wb.active
    ws.title = "Контакты"

    header_fill = PatternFill("solid", fgColor="06B6D4")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    rub_fmt = '#,##0" ₽";-#,##0" ₽"'
    int_fmt = '#,##0'

    headers = [
        "Телефон", "ФИО", "Визитов",
        "Первый визит", "Последний визит", "Дней с последнего визита",
        "Сумма (₽)", "LTV (₽)",
    ]
    widths = [16, 28, 10, 14, 16, 14, 16, 16]
    for col_idx, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = header_align
        c.border = border
        ws.column_dimensions[c.column_letter].width = w
    ws.row_dimensions[1].height = 22

    for i, item in enumerate(contacts, start=1):
        row = i + 1
        cells = [
            (1, item["phone"], None),
            (2, item["name"] or "—", None),
            (3, item["visits"], int_fmt),
            (4, item["first_visit"], None),
            (5, item["last_visit"], None),
            (6, item["days_since_last"], int_fmt),
            (7, round(item["total_spent"], 2), rub_fmt),
            (8, round(item["ltv"], 2), rub_fmt),
        ]
        for col_idx, val, fmt in cells:
            c = ws.cell(row=row, column=col_idx, value=val)
            c.border = border
            if fmt:
                c.number_format = fmt
            if col_idx in (3, 6):
                c.alignment = Alignment(horizontal="center")
            elif col_idx in (7, 8):
                c.alignment = Alignment(horizontal="right")
            elif col_idx in (4, 5):
                c.alignment = Alignment(horizontal="center")

    if not contacts:
        cell = ws.cell(row=2, column=1, value="Нет контактов под текущие фильтры.")
        cell.font = Font(italic=True, color="713F12")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.get("/contacts.csv", dependencies=[_mgr, _mod])
async def export_contacts(
    clinic_id: Optional[uuid.UUID] = Query(None),
    min_visits: int = Query(1, ge=1, description="Минимум визитов (1 = все, 2 = только повторные)"),
    inactive_days: Optional[int] = Query(
        None, ge=1, le=3650,
        description="Дней с последнего визита ≥ inactive_days (для «спящих»)",
    ),
    format: str = Query("csv", pattern="^(csv|xlsx)$", description="Формат: csv|xlsx"),
    user: User = Depends(require_manager),
    tenant: Tenant | None = Depends(get_current_tenant),
    db: AsyncSession = Depends(get_tenant_db),
):
    """Экспорт контактов пациентов (телефон, ФИО, визиты, даты, суммы)."""
    if tenant is None:
        raise HTTPException(status_code=400, detail="Тенант не определён")

    effective_clinic_id = await _resolve_clinic_scope(db, user, clinic_id)

    try:
        contacts = await _fetch_contacts(
            db, tenant, effective_clinic_id, min_visits, inactive_days,
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка выборки: {e}")

    today = datetime.utcnow().strftime("%Y-%m-%d")
    if format == "xlsx":
        try:
            data = _build_contacts_xlsx(contacts)
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Ошибка генерации XLSX: {e}")
        filename = f"LTV-контакты-АРЦ-{today}.xlsx"
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    else:
        data = _build_contacts_csv(contacts)
        filename = f"LTV-контакты-АРЦ-{today}.csv"
        media_type = "text/csv; charset=utf-8"

    return Response(
        content=data,
        media_type=media_type,
        headers={"Content-Disposition": _content_disposition(filename)},
    )
