"""
LTV-экспорт: PDF (WeasyPrint+Jinja2) и Excel (openpyxl).

Используется endpoint'ами:
  GET /analytics/ltv/export/pdf   → application/pdf
  GET /analytics/ltv/export/xlsx  → application/vnd.openxml...spreadsheet

Архитектура:
  generate_ltv_pdf(db, tenant, clinic_id, period) -> bytes
    - тянет summary/patients/cohorts через те же выборки, что и REST-endpoints
    - рендерит шаблон templates/ltv_report.html (Jinja2)
    - SVG-диаграммы строятся вручную (без matplotlib) — bar по когортам, pie по churn
    - возвращает PDF-байты, готовые к stream-отдаче

  generate_ltv_excel(db, tenant, clinic_id) -> bytes
    - openpyxl: 3 листа (Сводка, Топ пациентов, Когорты)
    - шапка с фоном #06b6d4, рамки, числа с разделителями, ₽-суффикс

Если данных нет (PatientLtvSnapshot пустой) — отдаём документ с пометкой
«Нет данных за период, запустите пересчёт».
"""
from __future__ import annotations

import io
import math
import uuid
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Optional

from jinja2 import Environment, FileSystemLoader, select_autoescape
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.clinic import Clinic
from app.models.ltv import PatientLtvSnapshot
from app.models.tenant import Tenant
from app.services.ltv_service import compute_cohorts


# Базовый горизонт, под который рассчитаны и сохранены снапшоты в БД.
# Если запросили другой — масштабируем коэффициентом years/3.
_BASE_LTV_HORIZON_YEARS = Decimal("3")


def _scale(value, years: int) -> float:
    """Пересчёт сохранённого LTV/NetLTV из базовых 3 лет в `years`."""
    if value is None:
        return 0.0
    try:
        d = Decimal(str(value))
    except Exception:
        return 0.0
    res = (d * Decimal(years) / _BASE_LTV_HORIZON_YEARS).quantize(Decimal("0.01"))
    return float(res)


def _years_label_ru(years: int) -> str:
    """Русское склонение «лет/года/год» для числа years."""
    n = int(years)
    if n % 10 == 1 and n % 100 != 11:
        return f"{n} год"
    if n % 10 in (2, 3, 4) and n % 100 not in (12, 13, 14):
        return f"{n} года"
    return f"{n} лет"

# ── Jinja2 окружение для PDF-шаблона ────────────────────────────────────────
_TEMPLATES_DIR = Path(__file__).resolve().parent.parent / "templates"
_jinja_env = Environment(
    loader=FileSystemLoader(str(_TEMPLATES_DIR)),
    autoescape=select_autoescape(["html", "xml"]),
)


# ───────────────────────────────────────────────────────────────────────────
# Хелперы форматирования
# ───────────────────────────────────────────────────────────────────────────

def _fmt_rub(v) -> str:
    """1 234 567 ₽ или «—» если ноль/None."""
    try:
        n = float(v or 0)
    except (TypeError, ValueError):
        return "—"
    if not math.isfinite(n) or n <= 0:
        return "—"
    return f"{int(round(n)):,}".replace(",", " ") + " ₽"


def _fmt_rub_zero_ok(v) -> str:
    """1 234 567 ₽ — но 0 тоже показываем как «0 ₽» (для KPI-таблиц)."""
    try:
        n = float(v or 0)
    except (TypeError, ValueError):
        return "0 ₽"
    if not math.isfinite(n):
        return "0 ₽"
    return f"{int(round(n)):,}".replace(",", " ") + " ₽"


def _fmt_int(v) -> str:
    try:
        n = int(v or 0)
    except (TypeError, ValueError):
        return "0"
    return f"{n:,}".replace(",", " ")


def _fmt_pct(v) -> str:
    try:
        n = float(v or 0)
    except (TypeError, ValueError):
        return "0%"
    if not math.isfinite(n):
        return "0%"
    return f"{n:.1f}%"


def _fmt_date_ru(dt) -> str:
    """ДД.ММ.ГГГГ или пусто (для PDF/CSV). dt — datetime или None."""
    if dt is None:
        return ""
    try:
        return dt.strftime("%d.%m.%Y")
    except Exception:
        return ""


# ───────────────────────────────────────────────────────────────────────────
# Сбор данных (общий для PDF и Excel)
# ───────────────────────────────────────────────────────────────────────────

async def _collect_ltv_data(
    db: AsyncSession,
    tenant: Tenant,
    clinic_id: Optional[uuid.UUID] = None,
    years: int = 3,
):
    """Возвращает кортеж (summary, patients, cohorts) — те же данные, что отдают
    REST-endpoints /analytics/ltv/{summary,patients,cohorts}.

    `years` — горизонт расчёта LTV (1..10). LTV/NetLTV пересчитываются из
    базового горизонта 3 года, сохранённого в БД, коэффициентом years/3.
    """
    # ── Summary ───────────────────────────────────────────────────────────
    base = select(
        func.count(PatientLtvSnapshot.id),
        func.coalesce(func.avg(PatientLtvSnapshot.ltv_estimate), 0),
        func.coalesce(func.sum(PatientLtvSnapshot.total_spent), 0),
        func.coalesce(func.avg(PatientLtvSnapshot.avg_check), 0),
        func.max(PatientLtvSnapshot.computed_at),
        func.coalesce(
            func.avg(func.nullif(PatientLtvSnapshot.net_ltv, 0)),
            0,
        ),
    ).where(PatientLtvSnapshot.tenant_id == tenant.id)
    if clinic_id is not None:
        base = base.where(PatientLtvSnapshot.clinic_id == clinic_id)
    row = (await db.execute(base)).one()
    total_patients, avg_ltv, total_spent, avg_check, last_computed, avg_net_ltv = row

    at_risk_q = select(func.count(PatientLtvSnapshot.id)).where(
        PatientLtvSnapshot.tenant_id == tenant.id,
        PatientLtvSnapshot.churn_risk == "high",
    )
    medium_q = select(func.count(PatientLtvSnapshot.id)).where(
        PatientLtvSnapshot.tenant_id == tenant.id,
        PatientLtvSnapshot.churn_risk == "medium",
    )
    low_q = select(func.count(PatientLtvSnapshot.id)).where(
        PatientLtvSnapshot.tenant_id == tenant.id,
        PatientLtvSnapshot.churn_risk == "low",
    )
    if clinic_id is not None:
        at_risk_q = at_risk_q.where(PatientLtvSnapshot.clinic_id == clinic_id)
        medium_q = medium_q.where(PatientLtvSnapshot.clinic_id == clinic_id)
        low_q = low_q.where(PatientLtvSnapshot.clinic_id == clinic_id)
    at_risk = int((await db.execute(at_risk_q)).scalar() or 0)
    medium = int((await db.execute(medium_q)).scalar() or 0)
    low = int((await db.execute(low_q)).scalar() or 0)

    total = int(total_patients or 0)
    churn_rate = round(((at_risk + medium) / total * 100.0), 2) if total else 0.0

    summary = {
        "total_patients": total,
        # avg_ltv / avg_net_ltv пересчитаны под горизонт `years`.
        "avg_ltv": _scale(avg_ltv, years),
        "avg_net_ltv": _scale(avg_net_ltv, years),
        "total_spent": float(total_spent or 0),
        "avg_check": float(avg_check or 0),
        "churn_rate": churn_rate,
        "at_risk_patients": at_risk,
        "medium_risk_patients": medium,
        "low_risk_patients": low,
        "last_computed_at": last_computed.isoformat() if last_computed else None,
        "horizon_years": years,
    }

    # ── Топ пациентов (top-50) ────────────────────────────────────────────
    pq = select(PatientLtvSnapshot).where(
        PatientLtvSnapshot.tenant_id == tenant.id,
        PatientLtvSnapshot.visits_count >= 2,
    )
    if clinic_id is not None:
        pq = pq.where(PatientLtvSnapshot.clinic_id == clinic_id)
    pq = pq.order_by(PatientLtvSnapshot.ltv_estimate.desc()).limit(50)
    prows = (await db.execute(pq)).scalars().all()
    patients = [
        {
            "patient_phone": r.patient_phone,
            "patient_name": r.patient_name,
            "visits_count": int(r.visits_count or 0),
            "total_spent": float(r.total_spent or 0),
            "avg_check": float(r.avg_check or 0),
            # ltv_estimate / net_ltv пересчитаны под горизонт `years`.
            "ltv_estimate": _scale(r.ltv_estimate, years),
            "net_ltv": _scale(r.net_ltv, years),
            "churn_risk": r.churn_risk or "high",
            "first_visit_at": r.first_visit_at,
            "last_visit_at": r.last_visit_at,
        }
        for r in prows
    ]

    # ── Когорты ───────────────────────────────────────────────────────────
    raw_cohorts = await compute_cohorts(db, tenant.id, period="quarter")
    # avg_ltv / avg_net_ltv в когортах тоже пересчитываем под выбранный горизонт.
    cohorts = [
        {
            **c,
            "avg_ltv": _scale(c.get("avg_ltv"), years),
            "avg_net_ltv": _scale(c.get("avg_net_ltv"), years),
        }
        for c in raw_cohorts
    ]

    return summary, patients, cohorts


# ───────────────────────────────────────────────────────────────────────────
# SVG-диаграммы (рисуем сами, без matplotlib)
# ───────────────────────────────────────────────────────────────────────────

def _bar_chart_svg(cohorts: list[dict], width: int = 240, height: int = 130) -> str:
    """Bar chart по когортам — высота столбика пропорциональна avg_ltv."""
    if not cohorts:
        return f'<svg width="{width}" height="{height}"></svg>'
    # Берём последние 8 когорт (визуально читаемо)
    items = list(reversed(cohorts[:8]))
    max_v = max((float(c.get("avg_ltv") or 0) for c in items), default=0) or 1.0
    pad_l, pad_r, pad_t, pad_b = 28, 8, 10, 24
    chart_w = width - pad_l - pad_r
    chart_h = height - pad_t - pad_b
    n = len(items)
    bar_w = chart_w / max(n, 1) * 0.7
    gap = chart_w / max(n, 1) * 0.3
    parts = [f'<svg width="{width}" height="{height}" viewBox="0 0 {width} {height}" xmlns="http://www.w3.org/2000/svg">']
    # Оси
    parts.append(f'<line x1="{pad_l}" y1="{pad_t}" x2="{pad_l}" y2="{pad_t+chart_h}" stroke="#cbd5e1" stroke-width="0.5"/>')
    parts.append(f'<line x1="{pad_l}" y1="{pad_t+chart_h}" x2="{pad_l+chart_w}" y2="{pad_t+chart_h}" stroke="#cbd5e1" stroke-width="0.5"/>')
    # Метка max
    parts.append(
        f'<text x="{pad_l-3}" y="{pad_t+8}" font-size="7" fill="#64748b" text-anchor="end">'
        f'{int(round(max_v)):,}'.replace(",", " ") + '</text>'
    )
    parts.append(
        f'<text x="{pad_l-3}" y="{pad_t+chart_h}" font-size="7" fill="#64748b" text-anchor="end">0</text>'
    )
    for i, c in enumerate(items):
        v = float(c.get("avg_ltv") or 0)
        h = (v / max_v) * chart_h if max_v > 0 else 0
        x = pad_l + i * (bar_w + gap) + gap / 2
        y = pad_t + chart_h - h
        parts.append(
            f'<rect x="{x:.1f}" y="{y:.1f}" width="{bar_w:.1f}" height="{h:.1f}" '
            f'fill="#06b6d4" rx="1"/>'
        )
        # Подпись когорты под столбиком
        label = str(c.get("cohort") or "")
        parts.append(
            f'<text x="{x + bar_w/2:.1f}" y="{pad_t+chart_h+10}" font-size="6.5" '
            f'fill="#475569" text-anchor="middle">{label}</text>'
        )
    parts.append('</svg>')
    return "".join(parts)


def _pie_chart_svg(low: int, medium: int, high: int, size: int = 130) -> str:
    """Pie по распределению churn-риска (low/medium/high)."""
    total = low + medium + high
    cx, cy, r = size / 2, size / 2, size / 2 - 6
    if total <= 0:
        return (
            f'<svg width="{size}" height="{size}" viewBox="0 0 {size} {size}" '
            f'xmlns="http://www.w3.org/2000/svg">'
            f'<circle cx="{cx}" cy="{cy}" r="{r}" fill="#e2e8f0"/></svg>'
        )
    parts = [
        f'<svg width="{size}" height="{size}" viewBox="0 0 {size} {size}" '
        f'xmlns="http://www.w3.org/2000/svg">'
    ]
    segments = [
        ("#10b981", low),
        ("#f59e0b", medium),
        ("#ef4444", high),
    ]
    # Если только один сегмент → рисуем целый круг, иначе path-арки
    nonzero = [s for s in segments if s[1] > 0]
    if len(nonzero) == 1:
        parts.append(f'<circle cx="{cx}" cy="{cy}" r="{r}" fill="{nonzero[0][0]}"/>')
    else:
        start = -math.pi / 2  # начинаем сверху
        for color, val in segments:
            if val <= 0:
                continue
            frac = val / total
            angle = frac * 2 * math.pi
            end = start + angle
            x1 = cx + r * math.cos(start)
            y1 = cy + r * math.sin(start)
            x2 = cx + r * math.cos(end)
            y2 = cy + r * math.sin(end)
            large = 1 if angle > math.pi else 0
            d = f"M{cx:.2f},{cy:.2f} L{x1:.2f},{y1:.2f} A{r:.2f},{r:.2f} 0 {large} 1 {x2:.2f},{y2:.2f} Z"
            parts.append(f'<path d="{d}" fill="{color}"/>')
            start = end
    # Метка по центру: total
    parts.append(
        f'<text x="{cx}" y="{cy+3}" font-size="9" fill="#fff" '
        f'font-weight="700" text-anchor="middle">{total}</text>'
    )
    parts.append('</svg>')
    return "".join(parts)


# ───────────────────────────────────────────────────────────────────────────
# PDF
# ───────────────────────────────────────────────────────────────────────────

async def generate_ltv_pdf(
    db: AsyncSession,
    tenant: Tenant,
    clinic_id: Optional[uuid.UUID] = None,
    period: str = "all",
    years: int = 3,
) -> bytes:
    """Сгенерировать PDF-отчёт LTV. Возвращает байты PDF.

    `years` — горизонт расчёта LTV (1..10 лет). Все суммы LTV/NetLTV
    пересчитываются под этот горизонт; шапка отчёта подписывается «N лет».
    """
    # Ленивый импорт WeasyPrint (тяжёлая зависимость)
    from weasyprint import HTML  # noqa: WPS433

    # Клампим years на всякий случай (если позвали из не-роутера)
    try:
        years = max(1, min(10, int(years)))
    except (TypeError, ValueError):
        years = 3

    summary, patients, cohorts = await _collect_ltv_data(db, tenant, clinic_id, years=years)

    # Лейбл клиники (если задана)
    clinic_label = "Все клиники"
    if clinic_id is not None:
        cr = await db.execute(select(Clinic).where(Clinic.id == clinic_id))
        cl = cr.scalar_one_or_none()
        if cl:
            clinic_label = cl.name or str(clinic_id)
        else:
            clinic_label = str(clinic_id)

    period_label = {
        "all": "Всё время",
        "month": "Месяц",
        "quarter": "Квартал",
        "year": "Год",
    }.get(period, period)

    has_data = (summary.get("total_patients") or 0) > 0

    # KPI-форматирование
    kpi = {
        "avg_ltv_fmt": _fmt_rub_zero_ok(summary["avg_ltv"]),
        "total_patients_fmt": _fmt_int(summary["total_patients"]),
        "churn_rate_fmt": _fmt_pct(summary["churn_rate"]),
        "at_risk_fmt": _fmt_int(summary["at_risk_patients"]),
    }

    # Форматируем пациентов
    patients_fmt = []
    for p in patients:
        patients_fmt.append({
            **p,
            "avg_check_fmt": _fmt_rub_zero_ok(p["avg_check"]),
            "total_spent_fmt": _fmt_rub_zero_ok(p["total_spent"]),
            "ltv_estimate_fmt": _fmt_rub_zero_ok(p["ltv_estimate"]),
            "net_ltv_fmt": _fmt_rub(p["net_ltv"]),  # 0 → «—»
            "first_visit_fmt": _fmt_date_ru(p.get("first_visit_at")),
            "last_visit_fmt": _fmt_date_ru(p.get("last_visit_at")),
        })

    cohorts_fmt = []
    for c in cohorts:
        cohorts_fmt.append({
            **c,
            "patients_fmt": _fmt_int(c.get("patients")),
            "total_spent_fmt": _fmt_rub_zero_ok(c.get("total_spent")),
            "avg_ltv_fmt": _fmt_rub_zero_ok(c.get("avg_ltv")),
            "avg_net_ltv_fmt": _fmt_rub(c.get("avg_net_ltv")),
        })

    # Диаграммы
    bar_svg = _bar_chart_svg(cohorts)
    pie_svg = _pie_chart_svg(
        summary.get("low_risk_patients", 0),
        summary.get("medium_risk_patients", 0),
        summary.get("at_risk_patients", 0),
    )

    ctx = {
        "tenant_name": "АРЦ КлиникСеть",
        "clinic_label": clinic_label,
        "period_label": period_label,
        "generated_at": datetime.utcnow().strftime("%d.%m.%Y %H:%M UTC"),
        "kpi": kpi,
        "patients": patients_fmt,
        "cohorts": cohorts_fmt,
        "has_data": has_data,
        "bar_chart_svg": bar_svg,
        "pie_chart_svg": pie_svg,
        # Горизонт LTV — для шапки отчёта и описаний секций
        "horizon_years": years,
        "horizon_label": _years_label_ru(years),
    }

    template = _jinja_env.get_template("ltv_report.html")
    html_str = template.render(**ctx)
    return HTML(string=html_str).write_pdf()


# ───────────────────────────────────────────────────────────────────────────
# Excel
# ───────────────────────────────────────────────────────────────────────────

async def generate_ltv_excel(
    db: AsyncSession,
    tenant: Tenant,
    clinic_id: Optional[uuid.UUID] = None,
    years: int = 3,
) -> bytes:
    """Сгенерировать Excel-отчёт LTV. Возвращает байты XLSX.

    `years` — горизонт расчёта LTV (1..10 лет). LTV/NetLTV пересчитываются
    под этот горизонт; в шапке листа «Сводка» — отдельная строка.
    """
    # Ленивый импорт openpyxl
    from openpyxl import Workbook  # noqa: WPS433
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # noqa: WPS433

    try:
        years = max(1, min(10, int(years)))
    except (TypeError, ValueError):
        years = 3

    summary, patients, cohorts = await _collect_ltv_data(db, tenant, clinic_id, years=years)
    has_data = (summary.get("total_patients") or 0) > 0
    horizon_label = _years_label_ru(years)

    # ── Стили ─────────────────────────────────────────────────────────────
    header_fill = PatternFill("solid", fgColor="06B6D4")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    rub_fmt = '#,##0" ₽";-#,##0" ₽"'
    pct_fmt = '0.00"%"'
    int_fmt = '#,##0'

    wb = Workbook()

    # ── Лист «Сводка» ─────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Сводка"
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 22

    rows_summary = [
        ("Метрика", "Значение"),
        ("Горизонт LTV / NetLTV", horizon_label),
        ("Всего пациентов", summary["total_patients"], int_fmt),
        (f"Avg LTV ({horizon_label})", round(summary["avg_ltv"], 2), rub_fmt),
        (f"Avg NetLTV ({horizon_label})", round(summary["avg_net_ltv"], 2), rub_fmt),
        ("Total spent", round(summary["total_spent"], 2), rub_fmt),
        ("Avg check", round(summary["avg_check"], 2), rub_fmt),
        ("Churn rate", round(summary["churn_rate"], 2), pct_fmt),
        ("Пациенты high-risk (at-risk)", summary["at_risk_patients"], int_fmt),
        ("Пациенты medium-risk", summary["medium_risk_patients"], int_fmt),
        ("Пациенты low-risk", summary["low_risk_patients"], int_fmt),
        ("Дата последнего пересчёта", summary.get("last_computed_at") or "—"),
    ]

    # Заголовок
    ws.cell(row=1, column=1, value=rows_summary[0][0])
    ws.cell(row=1, column=2, value=rows_summary[0][1])
    for col in (1, 2):
        c = ws.cell(row=1, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = header_align
        c.border = border

    for i, item in enumerate(rows_summary[1:], start=2):
        label = item[0]
        value = item[1]
        fmt = item[2] if len(item) > 2 else None
        a = ws.cell(row=i, column=1, value=label)
        b = ws.cell(row=i, column=2, value=value)
        a.border = border
        b.border = border
        a.alignment = Alignment(vertical="center")
        b.alignment = Alignment(horizontal="right", vertical="center")
        if fmt:
            b.number_format = fmt

    if not has_data:
        msg_row = len(rows_summary) + 2
        cell = ws.cell(
            row=msg_row, column=1,
            value="Нет данных за период, запустите пересчёт.",
        )
        cell.font = Font(italic=True, color="713F12")
        ws.merge_cells(start_row=msg_row, start_column=1, end_row=msg_row, end_column=2)

    # ── Лист «Топ пациентов» ──────────────────────────────────────────────
    ws_p = wb.create_sheet("Топ пациентов")
    headers_p = [
        "#", "Имя", "Телефон", "Визитов",
        "Первый визит", "Последний визит",
        "Ср. чек", "Total spent", "LTV", "NetLTV", "Churn risk",
    ]
    widths_p = [5, 28, 18, 10, 14, 16, 14, 16, 16, 16, 14]
    for col_idx, (h, w) in enumerate(zip(headers_p, widths_p), start=1):
        c = ws_p.cell(row=1, column=col_idx, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = header_align
        c.border = border
        ws_p.column_dimensions[c.column_letter].width = w
    ws_p.row_dimensions[1].height = 22

    if patients:
        for i, p in enumerate(patients, start=1):
            row = i + 1
            cells = [
                (1, i, int_fmt),
                (2, p.get("patient_name") or "—", None),
                (3, p.get("patient_phone") or "—", None),
                (4, int(p.get("visits_count") or 0), int_fmt),
                (5, _fmt_date_ru(p.get("first_visit_at")) or "—", None),
                (6, _fmt_date_ru(p.get("last_visit_at")) or "—", None),
                (7, round(float(p.get("avg_check") or 0), 2), rub_fmt),
                (8, round(float(p.get("total_spent") or 0), 2), rub_fmt),
                (9, round(float(p.get("ltv_estimate") or 0), 2), rub_fmt),
                (10, round(float(p.get("net_ltv") or 0), 2), rub_fmt),
                (11, p.get("churn_risk") or "—", None),
            ]
            for col_idx, val, fmt in cells:
                c = ws_p.cell(row=row, column=col_idx, value=val)
                c.border = border
                if fmt:
                    c.number_format = fmt
                if col_idx in (1, 4, 5, 6, 11):
                    c.alignment = Alignment(horizontal="center")
                elif col_idx in (7, 8, 9, 10):
                    c.alignment = Alignment(horizontal="right")
    else:
        cell = ws_p.cell(
            row=2, column=1,
            value="Нет пациентов с ≥ 2 визитами. Запустите пересчёт.",
        )
        cell.font = Font(italic=True, color="713F12")
        ws_p.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers_p))

    ws_p.freeze_panes = "A2"

    # ── Лист «Когорты» ────────────────────────────────────────────────────
    ws_c = wb.create_sheet("Когорты")
    headers_c = ["Квартал", "Пациентов", "Total spent", "LTV avg", "NetLTV avg"]
    widths_c = [15, 15, 20, 20, 20]
    for col_idx, (h, w) in enumerate(zip(headers_c, widths_c), start=1):
        c = ws_c.cell(row=1, column=col_idx, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = header_align
        c.border = border
        ws_c.column_dimensions[c.column_letter].width = w
    ws_c.row_dimensions[1].height = 22

    if cohorts:
        for i, co in enumerate(cohorts, start=1):
            row = i + 1
            cells = [
                (1, co.get("cohort") or "—", None),
                (2, int(co.get("patients") or 0), int_fmt),
                (3, round(float(co.get("total_spent") or 0), 2), rub_fmt),
                (4, round(float(co.get("avg_ltv") or 0), 2), rub_fmt),
                (5, round(float(co.get("avg_net_ltv") or 0), 2), rub_fmt),
            ]
            for col_idx, val, fmt in cells:
                c = ws_c.cell(row=row, column=col_idx, value=val)
                c.border = border
                if fmt:
                    c.number_format = fmt
                if col_idx == 1:
                    c.alignment = Alignment(horizontal="center")
                elif col_idx == 2:
                    c.alignment = Alignment(horizontal="center")
                else:
                    c.alignment = Alignment(horizontal="right")
    else:
        cell = ws_c.cell(
            row=2, column=1,
            value="Нет когортных данных. Запустите пересчёт.",
        )
        cell.font = Font(italic=True, color="713F12")
        ws_c.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers_c))

    ws_c.freeze_panes = "A2"

    # ── Сериализуем в bytes ───────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
